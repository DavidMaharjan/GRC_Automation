import os
import csv
import pytest
from playwright.sync_api import sync_playwright
from dotenv import load_dotenv

load_dotenv()

LOGIN_EMAIL    = os.getenv("LOGIN_EMAIL")
LOGIN_PASSWORD = os.getenv("LOGIN_PASSWORD")
APP_URL        = os.getenv("APP_URL")



# This list stores all test results. Each test adds one entry to it.
_results = []


def short_error(e):
    """Returns a short one-line error message like: AssertionError: Locator expected to be visible"""
    return f"{type(e).__name__}: {str(e).splitlines()[0]}"


def record_result(id, scenario, description, test_data, steps, expected, status, remarks):
    """Call this at the end of every test to save the result."""
    _results.append({
        "Test Case ID":            id,
        "Test Scenario":           scenario,
        "Test Case Description":   description,
        "Test Data":               test_data,
        "Steps":                   steps,
        "Expected Result":         expected,
        "Test Status (PASS/FAIL)": status,
        "Remarks":                 remarks,
    })



# Playwright fixtures — these set up the browser and pages for your tests

@pytest.fixture(scope="session")
def browser():
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        yield browser
        browser.close()


@pytest.fixture
def page(browser):
    page = browser.new_page()
    yield page
    page.close()


@pytest.fixture
def login(page):
    page.goto(f"{APP_URL}/sign-in")
    page.get_by_role("textbox", name="Enter your email (eg, john@").fill(LOGIN_EMAIL)
    page.get_by_role("textbox", name="Enter your password").fill(LOGIN_PASSWORD)
    page.get_by_role("button", name="Sign In").click()
    page.wait_for_timeout(5000)



# After ALL tests finish, write the Excel and CSV reports


HEADERS = [
    "Test Case ID", "Test Scenario", "Test Case Description", "Test Data",
    "Steps", "Expected Result", "Test Status (PASS/FAIL)", "Remarks",
]


def pytest_sessionfinish(session, exitstatus):
    if not _results:
        return

    os.makedirs("reports", exist_ok=True)

    # Find the next version number — e.g. V1, V2, V3 ...
    version = 1
    while os.path.isfile(os.path.join("reports", f"test_report_V{version}.xlsx")):
        version += 1

    # ---- CSV 
    csv_path = os.path.join("reports", "test_results.csv")
    file_exists = os.path.isfile(csv_path)
    with open(csv_path, "a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=HEADERS)
        if not file_exists:
            writer.writeheader()
        writer.writerows(_results)
    print(f"\n[Report] CSV  -> {csv_path}")

    # ---- Excel  
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test Results"

        # Style the header row
        header_fill = PatternFill("solid", fgColor="2E4057")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin_border = Border(
            left=Side(style="thin", color="AAAAAA"),
            right=Side(style="thin", color="AAAAAA"),
            top=Side(style="thin", color="AAAAAA"),
            bottom=Side(style="thin", color="AAAAAA"),
        )

        for col, header in enumerate(HEADERS, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

        # Colour rows green/red/yellow based on status
        pass_fill = PatternFill("solid", fgColor="D6F5D6")
        fail_fill = PatternFill("solid", fgColor="FFD6D6")
        skip_fill = PatternFill("solid", fgColor="FFF3CD")

        for row, result in enumerate(_results, start=2):
            status = result.get("Status", "")
            if status == "PASS":
                row_fill = pass_fill
            elif status == "FAIL":
                row_fill = fail_fill
            else:
                row_fill = skip_fill

            for col, key in enumerate(HEADERS, start=1):
                cell = ws.cell(row=row, column=col, value=result.get(key, ""))
                cell.fill = row_fill
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = thin_border

        # Set column widths
        col_widths = [14, 20, 35, 35, 45, 50, 20, 50]
        for col, width in enumerate(col_widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        ws.row_dimensions[1].height = 30

        xlsx_path = os.path.join("reports", f"test_report_V{version}.xlsx")
        wb.save(xlsx_path)
        print(f"[Report] Excel -> {xlsx_path}")

    except ImportError:
        print("[Report] openpyxl not installed. Run: pip install openpyxl")
